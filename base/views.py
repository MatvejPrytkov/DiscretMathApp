import os

from django.template.loader import render_to_string
import json
from .utils import notify_teacher_about_submission, notify_teacher_about_test_completion, notify_students_about_new_lab, notify_students_about_new_test, notify_student_about_lab_grade, notify_student_about_test_grade, notify_student_about_new_test
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib import messages
from .forms import RegistrationForm, GradeTestForm, ProfileEditForm
from .models import LabWork, LabSubmission, TestResult, TestKindCategory, TestKindConfig, Notification, TeacherPersonalQuestion, TeacherTestPersonalQuestion
from django.contrib.auth.forms import AuthenticationForm
import pandas as pd
import random
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from .models import TestAnswer
from .forms import UserUpdateForm, PasswordChangeForm, TeacherTestForm, TestQuestionForm, TeacherTestWithPersonalForm, TeacherPersonalQuestionForm
from .decorators import student_required, teacher_required, check_test_result_access, any_user_required
from django.db.models import Count, Sum, Avg, Max, CharField, Case, When, FloatField, Value
from django.db.models.functions import Round, Cast
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from .forms import MyLoginForm, LabWorkForm, AddQuestionForm, CreateTeacherTestForm
from django.db.models import Q
from django.core.paginator import Paginator
from .models import TestCategory, TestQuestion, TeacherTest, TeacherTestQuestion, UserProfile, Message, TeacherStudentMessage
from django.forms import formset_factory
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from io import BytesIO
from .chat_views import groupmates_list, chat_detail, send_message, get_unread_messages_count, get_chat_users_list
from .chat_teacher_views import (
    student_teacher_chat_list,
    teacher_student_chat_list,
    teacher_student_chat_detail,
    send_teacher_student_message,
    get_teacher_student_messages
)
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .context_processors import has_final_test
# Добавьте эту функцию в views.py после импортов


def search_teachers_api(request):
    """API для поиска преподавателей по ФИО (для регистрации)"""
    query = request.GET.get('q', '').strip()

    # Базовый запрос - все пользователи с ролью teacher
    teachers = User.objects.filter(profile__role='teacher').select_related('profile')

    if query:
        # Поиск по ФИО (full_name), email и username
        teachers = teachers.filter(
            Q(profile__full_name__icontains=query) |
            Q(email__icontains=query) |
            Q(username__icontains=query)
        )

    # Ограничиваем количество результатов
    teachers = teachers[:10]

    # Формируем ответ
    teachers_data = []
    for teacher in teachers:
        full_name = ''
        if hasattr(teacher, 'profile') and teacher.profile:
            full_name = teacher.profile.full_name
        if not full_name:
            full_name = teacher.get_full_name() or teacher.username

        teachers_data.append({
            'id': teacher.id,
            'full_name': full_name,
            'email': teacher.email,
            'username': teacher.username
        })

    return JsonResponse({
        'success': True,
        'teachers': teachers_data,
        'query': query
    })

def get_file_response(file_field, filename):
    """Возвращает правильный HTTP ответ для файла в зависимости от его типа"""
    file_path = file_field.path
    file_extension = os.path.splitext(filename)[1].lower()

    # MIME-типы для разных форматов
    mime_types = {
        '.pdf': 'application/pdf',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.txt': 'text/plain',
        '.doc': 'application/msword',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xls': 'application/vnd.ms-excel',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.ppt': 'application/vnd.ms-powerpoint',
        '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
        '.zip': 'application/zip',
        '.rar': 'application/x-rar-compressed',
        '.7z': 'application/x-7z-compressed',
        '.mp4': 'video/mp4',
        '.mp3': 'audio/mpeg',
    }

    content_type = mime_types.get(file_extension, 'application/octet-stream')

    # Для файлов, которые можно просмотреть в браузере (PDF, изображения, текстовые)
    viewable_extensions = {'.pdf', '.jpg', '.jpeg', '.png', '.gif', '.txt'}

    if file_extension in viewable_extensions:
        # Открываем в браузере
        response = HttpResponse(content_type=content_type)
        with open(file_path, 'rb') as f:
            response.write(f.read())
        response['Content-Disposition'] = f'inline; filename="{filename}"'
    else:
        # Скачиваем файл
        response = HttpResponse(content_type=content_type)
        with open(file_path, 'rb') as f:
            response.write(f.read())
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
def index(request):
    return render(request, 'index.html')


# views.py - полная функция test_view с исправлениями

def _get_answer_text(question, option_letter):
    """Получает текст ответа по букве варианта"""
    if not option_letter:
        return ''
    mapping = {
        'a': question.option_a,
        'b': question.option_b,
        'c': question.option_c,
        'd': question.option_d,
    }
    return mapping.get(option_letter.lower(), '')


@login_required
@student_required
def test_view(request, test_kind: str):
    # Получаем конфигурацию теста из БД
    try:
        test_config = TestKindConfig.objects.get(code=test_kind, is_active=True)
    except TestKindConfig.DoesNotExist:
        messages.error(request, f"Тест '{test_kind}' не найден или неактивен")
        return redirect('profile')

    # ========== ПРОВЕРКА ДЛЯ ИТОГОВОГО ТЕСТА ==========
    if test_kind == 'final':
        existing_result = TestResult.objects.filter(
            user=request.user,
            test_type='final'
        ).exists()

        if existing_result:
            messages.warning(request, 'Вы уже прошли итоговое тестирование. Повторное прохождение невозможно.')
            final_result = TestResult.objects.filter(user=request.user, test_type='final').first()
            if final_result:
                return redirect('result_detail', pk=final_result.id)
            return redirect('profile')

    # =============== POST (проверка и сохранение) ===============
    if request.method == 'POST':
        results = {}
        total_correct = 0
        all_question_ids = []

        # Получаем все категории для этого типа теста
        test_categories = test_config.categories.all()

        if not test_categories.exists():
            messages.error(request, 'Тест не настроен. Обратитесь к администратору.')
            return redirect('profile')

        # ВАЖНО: Собираем ID вопросов из скрытых полей (не из radio-кнопок!)
        for key in request.POST.keys():
            if key.startswith('question_id_'):
                # Извлекаем ID вопроса из ключа вида "question_id_category_id"
                q_id = request.POST.get(key)
                if q_id and q_id.isdigit():
                    all_question_ids.append(int(q_id))

        # Удаляем дубликаты
        all_question_ids = list(set(all_question_ids))
        total_questions_all = len(all_question_ids)

        if total_questions_all == 0:
            messages.error(request, 'В тесте нет вопросов. Обратитесь к администратору.')
            return redirect('profile')

        # Обрабатываем ответы по категориям
        for category in test_categories:
            try:
                kind_category = TestKindCategory.objects.get(
                    test_kind=test_config,
                    category=category
                )
                questions_needed = kind_category.questions_count

                # Получаем ID вопросов для этой категории
                prefix = f'question_id_{category.code}_'
                category_question_ids = []
                for key in request.POST.keys():
                    if key.startswith(prefix):
                        q_id = request.POST.get(key)
                        if q_id and q_id.isdigit():
                            category_question_ids.append(int(q_id))

                category_question_ids = list(set(category_question_ids))

                if not category_question_ids:
                    continue

                correct_count = 0
                question_details = []

                for q_id in category_question_ids:
                    # Получаем ответ пользователя
                    user_answer = request.POST.get(f'q_{category.code}_{q_id}')
                    if user_answer:
                        user_answer = user_answer.strip().lower()
                    # else: user_answer остается None

                    try:
                        question = TestQuestion.objects.get(
                            id=q_id,
                            category=category,
                            is_active=True
                        )
                        is_correct = (user_answer is not None and user_answer == question.correct_option)

                        if is_correct:
                            correct_count += 1

                        question_details.append({
                            'question_id': q_id,
                            'user_answer': user_answer,
                            'correct_answer': question.correct_option,
                            'is_correct': is_correct,
                            'question_text': question.question_text
                        })

                    except TestQuestion.DoesNotExist:
                        question_details.append({
                            'question_id': q_id,
                            'user_answer': user_answer,
                            'correct_answer': None,
                            'is_correct': False,
                            'question_text': 'Вопрос не найден'
                        })
                        continue

                results[category.code] = {
                    'name': category.name,
                    'correct': correct_count,
                    'total': len(category_question_ids),
                    'config_questions': questions_needed,
                    'questions': question_details
                }

                total_correct += correct_count

            except TestKindCategory.DoesNotExist:
                results[category.code] = {'name': category.name, 'correct': 0, 'total': 0}
            except Exception as e:
                results[category.code] = {'name': category.name, 'correct': 0, 'total': 0}

        # Сохраняем результат
        test_result = TestResult.objects.create(
            user=request.user,
            test_type=test_kind,
            score=total_correct,
            total_questions=total_questions_all,
            percent=round((total_correct / total_questions_all) * 100, 2) if total_questions_all > 0 else 0,
            correct_answers=total_correct,
            percentage=round((total_correct / total_questions_all) * 100, 2) if total_questions_all > 0 else 0,
            category_results=results
        )

        # Сохраняем детальные ответы
        for category_code, data in results.items():
            for question_detail in data.get('questions', []):
                if question_detail.get('correct_answer') is None:
                    continue

                try:
                    question = TestQuestion.objects.get(id=question_detail['question_id'])

                    TestAnswer.objects.create(
                        result=test_result,
                        question_id=question.id,
                        question_text=question.question_text,
                        user_answer=question_detail['user_answer'],
                        user_answer_text=_get_answer_text(question, question_detail['user_answer']),
                        correct_answer=question.correct_option,
                        correct_answer_text=_get_answer_text(question, question.correct_option),
                        is_correct=question_detail['is_correct']
                    )
                except TestQuestion.DoesNotExist:
                    continue

        # Уведомление учителю
        teacher = None
        if hasattr(request.user, 'profile') and request.user.profile.teacher:
            teacher = request.user.profile.teacher

        if teacher:
            test_name = "Входное тестирование" if test_kind == 'start' else "Итоговое тестирование"
            from .utils import create_notification
            create_notification(
                recipient=teacher,
                sender=request.user,
                notification_type='test_completed',
                title=f'📊 Пройден тест: {test_name}',
                message=f'Студент {request.user.profile.full_name} прошел {test_name}. Результат: {total_correct}/{total_questions_all} ({round((total_correct / total_questions_all) * 100, 2)}%).',
                link=f'/teacher/result/{test_result.id}/'
            )

        return render(request, test_config.result_template, {
            'result': test_result,
            'answers': test_result.answers.all().order_by('question_id'),
            'results_by_category': results,
            'total': total_correct,
            'total_questions': total_questions_all,
            'percent': test_result.percent,
            'test_title': test_config.title,
            'test_config': test_config,
            'back_url': 'profile',
            'back_text': 'Вернуться в профиль',
            'active_nav': 'final_result' if test_kind == 'final' else None,
        })

    # =============== GET (показ вопросов) ===============
    questions_to_render = []
    test_categories = test_config.categories.filter(is_active=True)

    for category in test_categories:
        try:
            kind_category = TestKindCategory.objects.get(
                test_kind=test_config,
                category=category
            )
            questions_needed = kind_category.questions_count

            if questions_needed == 0:
                continue

            queryset = TestQuestion.objects.filter(
                category=category,
                is_active=True
            )

            if not queryset.exists():
                continue

            if questions_needed > 0:
                questions = list(queryset.order_by('?')[:questions_needed])
            else:
                questions = list(queryset.order_by('?'))

            for question in questions:
                questions_to_render.append({
                    'category': category.code,
                    'id_val': question.id,
                    'question': question.question_text,
                    'answer1': question.option_a,
                    'answer2': question.option_b,
                    'answer3': question.option_c,
                    'answer4': question.option_d,
                    'correct_answer': question.correct_option,
                    'difficulty': question.get_difficulty_display()
                })

        except TestKindCategory.DoesNotExist:
            continue
        except TestQuestion.DoesNotExist:
            continue

    random.shuffle(questions_to_render)

    return render(request, test_config.template, {
        'questions': questions_to_render,
        'test_title': test_config.title,
        'test_kind': test_kind,
        'test_type': test_kind,
        'test_config': test_config,
        'categories': test_categories,
        'active_nav': 'final_test' if test_kind == 'final' else None,
    })

def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)

            full_name = getattr(getattr(user, 'profile', None), 'full_name', user.get_username())
            messages.success(request, f'Регистрация успешна! Добро пожаловать, {full_name}!')

            # Получаем роль пользователя из формы
            role = form.cleaned_data.get('role')

            # Перенаправляем в зависимости от роли
            if role == 'teacher':
                return redirect('teacher_dashboard')  # Используем имя маршрута из urls.py
            else:
                return redirect('profile')  # Для учеников

        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = RegistrationForm()

    return render(request, 'register.html', {'form': form})


def login_view(request):
    if request.method == "POST":
        # Используем MyLoginForm вместо AuthenticationForm
        form = MyLoginForm(data=request.POST)

        if form.is_valid():
            user = form.get_user()
            selected_role = request.POST.get("role")  # 'student' или 'teacher'
            profile = getattr(user, 'profile', None)
            actual_role = getattr(profile, 'role', None)

            # Если у пользователя нет профиля/роли — лучше запретить вход
            if not actual_role:
                messages.error(request, "У аккаунта не задана роль. Обратитесь к администратору.")
                return render(request, "login.html", {"form": form})

            # Проверка на несовпадение роли
            if selected_role != actual_role:
                messages.error(request, "Вы выбрали неверную роль для этого аккаунта.")
                return render(request, "login.html", {"form": form})
            login(request, user)
            if actual_role == 'teacher':
                return redirect('teacher_dashboard')
            return redirect('profile')


        # Если данные неверны, Django автоматически передаст форму с нашими русскими ошибками
        return render(request, "login.html", {"form": form})

    else:
        form = MyLoginForm()

    return render(request, "login.html", {"form": form})

def logout_view(request):
    logout(request)
    messages.success(request, 'Вы успешно вышли из системы.')
    return redirect('initial')


@login_required
@student_required
def profile(request):
    user = request.user

    # Получаем параметры фильтрации из GET-запроса
    type_filter = request.GET.get('type', '')
    grade_filter = request.GET.get('grade', '')
    status_filter = request.GET.get('status', '')
    anchor = request.GET.get('anchor', '')  # Добавляем получение якоря

    # Оценки за лабораторные
    lab_submissions = LabSubmission.objects.filter(
        student=user,
    ).select_related('lab_work')

    # Применяем фильтры к лабораторным
    if type_filter and type_filter != 'lab':
        lab_submissions = lab_submissions.none()

    if grade_filter:
        if grade_filter == 'pending':
            lab_submissions = lab_submissions.filter(Q(grade__isnull=True) | Q(grade=''))
        else:
            lab_submissions = lab_submissions.filter(grade=grade_filter)

    if status_filter:
        lab_submissions = lab_submissions.filter(status=status_filter)

    # Оценки за тесты
    test_grades = TestResult.objects.filter(
        user=user
    ).order_by('-date_completed')

    # Применяем фильтры к тестам
    if type_filter and type_filter != 'lab':
        if type_filter == 'start':
            test_grades = test_grades.filter(test_type='start')
        elif type_filter == 'final':
            test_grades = test_grades.filter(test_type='final')
        elif type_filter == 'teacher':
            test_grades = test_grades.filter(test_type='teacher')
    elif type_filter == 'lab':
        test_grades = test_grades.none()

    if grade_filter:
        if grade_filter == 'pending':
            test_grades = test_grades.filter(grade__isnull=True)
        else:
            test_grades = test_grades.filter(grade=grade_filter)

    if status_filter:
        if status_filter == 'graded':
            test_grades = test_grades.filter(grade__isnull=False)
        elif status_filter == 'pending':
            test_grades = test_grades.filter(grade__isnull=True)

    total_count = lab_submissions.count() + test_grades.count()

    teacher_tests = TeacherTest.objects.filter(assigned_to=user, is_active=True).distinct()

    # Получаем данные об учителе ученика
    teacher = None
    teacher_name = None
    if hasattr(user, 'profile') and user.profile.teacher:
        teacher = user.profile.teacher
        if hasattr(teacher, 'profile') and teacher.profile.full_name:
            teacher_name = teacher.profile.full_name
        else:
            teacher_name = teacher.get_full_name() or teacher.username

    # Получаем информацию о итоговом тесте через контекстный процессор
    final_test_context = has_final_test(request)

    # СОЗДАЁМ СЛОВАРЬ КОНТЕКСТА
    context = {
        'user': user,
        'lab_grades': lab_submissions,
        'test_grades': test_grades,
        'results': test_grades,
        'teacher_tests': teacher_tests,
        'total_count': total_count,
        'teacher': teacher,
        'teacher_name': teacher_name,
        # Передаем текущие фильтры в шаблон
        'current_type_filter': type_filter,
        'current_grade_filter': grade_filter,
        'current_status_filter': status_filter,
        'anchor': anchor,  # Передаем якорь в шаблон
        # ДОБАВЛЯЕМ ПЕРЕМЕННЫЕ ДЛЯ ИТОГОВОГО ТЕСТА
        'has_final_test': final_test_context.get('has_final_test', False),
        'final_test_result_obj': final_test_context.get('final_test_result', None),
    }

    return render(request, 'student/Profile.html', context)



@login_required
@check_test_result_access
def result_detail(request, pk):
    try:
        result = TestResult.objects.get(pk=pk)
    except TestResult.DoesNotExist:
        messages.error(request, "Результат теста не найден")
        if hasattr(request.user, 'profile') and request.user.profile.role == 'teacher':
            return redirect('teacher_dashboard')
        return redirect('profile')

    # Определяем роль пользователя
    is_teacher = getattr(request.user.profile, 'role', '') == 'teacher'
    is_student = getattr(request.user.profile, 'role', '') == 'student'

    # СТРОГАЯ ПРОВЕРКА ПРАВ ДОСТУПА
    has_access = False

    if is_student:
        if result.user == request.user:
            has_access = True
        else:
            messages.error(request, "Вы не можете просматривать результаты других учеников")
            return redirect('profile')

    elif is_teacher:
        if hasattr(result.user, 'profile') and result.user.profile.teacher == request.user:
            has_access = True
        else:
            messages.error(request, "Этот результат теста принадлежит не вашему ученику")
            return redirect('teacher_dashboard')

    if not has_access:
        messages.error(request, "У вас нет прав для просмотра этого результата")
        return redirect('teacher_dashboard' if is_teacher else 'profile')

    # Обработка формы оценки (только для учителей)
    if is_teacher and request.method == 'POST':
        form = GradeTestForm(request.POST, instance=result)
        if form.is_valid():
            saved_result = form.save(commit=False)
            saved_result.graded_by = request.user
            saved_result.graded_at = timezone.now()
            saved_result.save()
            notify_student_about_test_grade(request.user, result.user, saved_result)
            messages.success(request, 'Оценка успешно сохранена!')
            # ✅ ПРАВИЛЬНЫЙ РЕДИРЕКТ ДЛЯ УЧИТЕЛЯ
            return redirect('teacher_result_detail', pk=pk)
    else:
        form = GradeTestForm(instance=result)

    answers = result.answers.all().order_by('question_id')
    results_by_category = result.category_results

    # ✅ ВЫБИРАЕМ ШАБЛОН В ЗАВИСИМОСТИ ОТ РОЛИ
    if is_teacher:
        template_name = 'teacher/result_detail.html'
        back_url = 'student_results'
        back_text = 'Назад к результатам учеников'
    else:
        template_name = 'student/result_detail.html'
        back_url = 'profile'
        back_text = 'Назад в профиль'
    active_nav = None
    if is_teacher:
        active_nav = 'teacher'
    else:
        # Для студента определяем тип теста
        if result.test_type == 'final':
            active_nav = 'final_result'
        else:
            # Все остальные тесты (start и teacher) ведут на страницу результатов
            active_nav = 'test_results'
    return render(request, template_name, {
        'result': result,
        'answers': answers,
        'results_by_category': results_by_category,
        'form': form,
        'is_teacher': is_teacher,
        'back_url': back_url,
        'back_text': back_text,
        'active_nav': active_nav,
    })


@login_required
def profile_update(request):
    """Редактирование профиля пользователя"""

    # Получаем профиль пользователя
    user_profile = getattr(request.user, 'profile', None)
    is_teacher = user_profile and user_profile.role == 'teacher'
    is_student = user_profile and user_profile.role == 'student'

    if request.method == 'POST':
        if is_student:
            # Для учеников используем расширенную форму
            form = ProfileEditForm(request.POST, instance=request.user, profile=user_profile)
        else:
            # Для учителей только базовые поля
            form = UserUpdateForm(request.POST, instance=request.user)

        if form.is_valid():
            form.save()
            messages.success(request, 'Профиль успешно обновлён!')

            # Перенаправляем в зависимости от роли
            if is_teacher:
                return redirect('teacher_dashboard')
            return redirect('profile')
    else:
        if is_student:
            form = ProfileEditForm(instance=request.user, profile=user_profile)
        else:
            form = UserUpdateForm(instance=request.user)

    # Статистика для ученика (если нужно)
    context = {
        'form': form,
        'is_student': is_student,
        'is_teacher': is_teacher,
        'user_profile': user_profile,
    }

    # Для учеников получаем доступных преподавателей для отображения
    if is_student:
        context['teachers'] = User.objects.filter(profile__role='teacher').select_related('profile')

    return render(request, 'profile_update.html', context)
@login_required
@any_user_required
def change_password(request):
    """Смена пароля"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.POST)
        if form.is_valid():
            user = request.user
            if user.check_password(form.cleaned_data['old_password']):
                user.set_password(form.cleaned_data['new_password1'])
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Пароль успешно изменён!')
                # Перенаправляем в зависимости от роли
                if hasattr(request.user, 'profile') and request.user.profile.role == 'teacher':
                    return redirect('teacher_dashboard')
                return redirect('profile')
            else:
                form.add_error('old_password', 'Неверный текущий пароль')
    else:
        form = PasswordChangeForm()

    return render(request, 'change_password.html', {'form': form})

@login_required
@any_user_required
def delete_account(request):
    """Удаление аккаунта"""
    if request.method == 'POST':
        user = request.user
        username = user.username
        user.delete()
        messages.success(request, f'Аккаунт {username} удалён.')
        return redirect('initial')

    return render(request, 'delete_account.html', {'user': request.user})
@login_required
@teacher_required
def student_results(request):
    """Просмотр результатов: либо всех 'своих' учеников, либо конкретного"""
    student_id = request.GET.get('student_id')
    type_filter = request.GET.get('type_filter', '')
    grade_filter = request.GET.get('grade_filter', '')
    group_filter = request.GET.get('group_filter', '')
    course_filter = request.GET.get('course_filter', '')

    # Базовый фильтр: только ученики, привязанные к данному учителю
    students = User.objects.filter(
        profile__role='student',
        profile__teacher=request.user
    ).select_related('profile')

    # Получаем уникальные группы и курсы для выпадающих списков
    available_groups = sorted(set(
        student.profile.group for student in students
        if student.profile and student.profile.group
    ))
    available_courses = sorted(set(
        student.profile.course for student in students
        if student.profile and student.profile.course
    ))

    # Если выбран конкретный ученик - проверяем, что он принадлежит учителю
    selected_student = None
    if student_id:
        try:
            # Пытаемся найти ученика среди своих
            selected_student = students.get(id=student_id)
        except User.DoesNotExist:
            # Ученик не найден или не принадлежит учителю
            messages.error(request, "Выбранный ученик не найден или не принадлежит вам")
            # Очищаем параметр student_id
            student_id = None
            # Перенаправляем на ту же страницу без student_id
            return redirect('student_results')
        except User.MultipleObjectsReturned:
            messages.error(request, "Ошибка: найдено несколько учеников")
            student_id = None

    # Результаты тестов учеников
    test_results = TestResult.objects.filter(
        user__profile__teacher=request.user
    ).select_related('user', 'user__profile').order_by('-date_completed')

    # Результаты лабораторных работ
    lab_results = LabSubmission.objects.filter(
        student__profile__teacher=request.user
    ).select_related('student__profile', 'lab_work').order_by('-submitted_at')

    # Фильтрация по конкретному ученику
    if student_id:
        test_results = test_results.filter(user_id=student_id)
        lab_results = lab_results.filter(student_id=student_id)

    # Фильтрация по группе
    if group_filter:
        test_results = test_results.filter(user__profile__group=group_filter)
        lab_results = lab_results.filter(student__profile__group=group_filter)

    # Фильтрация по курсу
    if course_filter:
        test_results = test_results.filter(user__profile__course=course_filter)
        lab_results = lab_results.filter(student__profile__course=course_filter)

    # Применяем фильтр по типу
    if type_filter == 'test':
        lab_results = lab_results.none()
    elif type_filter == 'lab':
        test_results = test_results.none()
    elif type_filter in ['start', 'final', 'teacher']:
        # Фильтруем тесты по подтипу
        test_results = test_results.filter(test_type=type_filter)
        lab_results = lab_results.none()

    # Применяем фильтр по оценке
    if grade_filter:
        if grade_filter == 'pending':
            test_results = test_results.filter(grade__isnull=True)
            lab_results = lab_results.filter(grade__isnull=True)
        else:
            test_results = test_results.filter(grade=int(grade_filter))
            lab_results = lab_results.filter(grade=grade_filter)

    # Преобразуем лабораторные работы в формат, похожий на TestResult для единого отображения
    combined_results = []

    # Добавляем результаты тестов
    for result in test_results:
        profile = result.user.profile
        combined_results.append({
            'type': 'test',
            'test_subtype': result.test_type,
            'user': result.user,
            'user_name': profile.full_name or result.user.username,
            'title': result.get_test_type_display(),
            'score': f"{result.score}/{result.total_questions}",
            'percent': result.percent,
            'grade': result.get_grade_display() if result.grade else None,
            'grade_value': result.grade,
            'date': result.date_completed,
            'detail_url': f'/teacher/result/{result.id}/',
            'group': profile.group if profile and profile.group else '—',
            'course': profile.course if profile and profile.course else '—',
        })

    # Добавляем результаты лабораторных работ
    for submission in lab_results:
        profile = submission.student.profile
        grade_num = None
        if submission.grade == '5':
            grade_num = 5
        elif submission.grade == '4':
            grade_num = 4
        elif submission.grade == '3':
            grade_num = 3
        elif submission.grade == '2':
            grade_num = 2

        combined_results.append({
            'type': 'lab',
            'user': submission.student,
            'user_name': profile.full_name or submission.student.username,
            'title': f"Лабораторная: {submission.lab_work.title}",
            'score': submission.grade or "—",
            'percent': None,
            'grade': submission.grade,
            'grade_value': grade_num,
            'date': submission.submitted_at,
            'status': submission.get_status_display(),
            'detail_url': f'/teacher/lab/submission/{submission.id}/',
            'group': profile.group if profile and profile.group else '—',
            'course': profile.course if profile and profile.course else '—',
        })

    # Сортируем по дате (сначала новые)
    combined_results.sort(key=lambda x: x['date'], reverse=True)

    return render(request, 'teacher/student_results.html', {
        'results': combined_results,
        'students': students,
        'selected_student': selected_student,
        'student_id': student_id,
        'group_filter': group_filter,
        'course_filter': course_filter,
        'available_groups': available_groups,
        'available_courses': available_courses,
    })
@teacher_required
def manage_tests(request):
    """Управление тестами (добавление/редактирование вопросов)"""
    # Здесь можно добавить логику для работы с Excel файлами
    return render(request, 'teacher/manage_tests.html')


@login_required
@teacher_required
def teacher_dashboard(request):
    if request.user.profile.role != 'teacher':
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    from django.db.models import Avg, Count, Q, Value, IntegerField, Case, When, Sum
    from django.db.models.functions import Coalesce

    # Получаем всех учеников учителя
    my_students = User.objects.filter(
        profile__role='student',
        profile__teacher=request.user
    ).select_related('profile')

    # Общее количество лабораторных работ учителя
    total_labs_count_value = LabWork.objects.filter(created_by=request.user, is_active=True).count()

    # Аннотируем каждого ученика
    students_annotated = []
    for student in my_students:
        # Все пройденные тесты (все типы)
        all_test_results = TestResult.objects.filter(user=student)
        total_tests_passed = all_test_results.count()

        # Количество сданных лабораторных
        lab_submission_count = LabSubmission.objects.filter(
            student=student,
            lab_work__created_by=request.user,
            status__in=['submitted', 'under_review', 'graded']
        ).count()

        # Средний балл за все тесты
        test_scores = [r.grade for r in all_test_results if r.grade and r.grade != 0]
        avg_grade = round(sum(test_scores) / len(test_scores), 1) if test_scores else None

        # Общее количество доступных тестов (входной + итоговый + тесты учителя)
        # Входной и итоговый тесты считаются как 2 доступных теста (если еще не пройдены, но они всегда доступны)
        # Тесты от учителя, назначенные ученику
        teacher_tests_count = TeacherTest.objects.filter(
            assigned_to=student,
            is_active=True
        ).count()

        # Всего доступных тестов = входной(1) + итоговый(1) + тесты учителя
        total_all_tests_available = 2 + teacher_tests_count
        # Количество пройденных тестов от учителя
        teacher_tests_passed = TestResult.objects.filter(
            user=student,
            test_type='teacher'
        ).count()

        # Количество назначенных тестов от учителя
        teacher_tests_count = TeacherTest.objects.filter(
            assigned_to=student,
            is_active=True
        ).count()

        students_annotated.append({
            'id': student.id,
            'profile': student.profile,
            'email': student.email,
            'total_tests_passed': total_tests_passed,
            'teacher_tests_passed': teacher_tests_passed,
            'teacher_tests_count': teacher_tests_count,
            'lab_submission_count': lab_submission_count,
            'total_labs_count': total_labs_count_value,
            'total_all_tests_available': total_all_tests_available,
            'avg_grade': avg_grade,
        })

    # 📊 ДАННЫЕ ДЛЯ ГРАФИКОВ
    test_stats = TestResult.objects.filter(
        user__profile__teacher=request.user
    ).values('test_type').annotate(
        count=Count('id')
    )

    # Распределение оценок за ТЕСТЫ
    test_grade_counts = TestResult.objects.filter(
        user__profile__teacher=request.user
    ).values('grade').annotate(count=Count('id'))

    test_grade_5_count = 0
    test_grade_4_count = 0
    test_grade_3_count = 0
    test_grade_2_count = 0
    test_grade_pending_count = 0

    for item in test_grade_counts:
        if item['grade'] == 5:
            test_grade_5_count = item['count']
        elif item['grade'] == 4:
            test_grade_4_count = item['count']
        elif item['grade'] == 3:
            test_grade_3_count = item['count']
        elif item['grade'] == 2:
            test_grade_2_count = item['count']

    test_grade_pending_count = TestResult.objects.filter(
        user__profile__teacher=request.user,
        grade__isnull=True
    ).count()

    # Распределение оценок за ЛАБОРАТОРНЫЕ
    lab_submissions = LabSubmission.objects.filter(
        student__profile__teacher=request.user
    )

    lab_grade_5_count = lab_submissions.filter(grade='5').count()
    lab_grade_4_count = lab_submissions.filter(grade='4').count()
    lab_grade_3_count = lab_submissions.filter(grade='3').count()
    lab_grade_2_count = lab_submissions.filter(grade='2').count()
    lab_grade_pending_count = lab_submissions.filter(grade__isnull=True).count()

    # Статистика по темам
    theme_stats = {}
    TRANSLATIONS = {
        'graphs': 'Графы',
        'logic': 'Логика',
        'plenty': 'Множества',
        'final': 'Итоговый',
        'start': 'Входной',
        'teacher_test': 'Тест от учителя'
    }
    all_results = TestResult.objects.filter(user__profile__teacher=request.user)
    for result in all_results:
        if result.category_results:
            for theme, data in result.category_results.items():
                if theme not in theme_stats:
                    display_name = TRANSLATIONS.get(theme, theme)
                    theme_stats[theme] = {'correct': 0, 'total': 0, 'name': display_name}
                theme_stats[theme]['correct'] += data.get('correct', 0)
                theme_stats[theme]['total'] += data.get('total', 0)

    total_tests_passed_all = sum(s['total_tests_passed'] for s in students_annotated)
    total_labs_submitted_all = sum(s['lab_submission_count'] for s in students_annotated)

    lab_count = total_labs_count_value
    pending_submissions = LabSubmission.objects.filter(
        status='under_review',
        lab_work__created_by=request.user
    ).count()
    groups_list = list(set(
        student['profile'].group for student in students_annotated
        if student['profile'] and student['profile'].group
    ))

    return render(request, 'teacher/teacher_dashboard.html', {
        'students': students_annotated,
        'students_count': len(students_annotated),
        'total_tests_passed': total_tests_passed_all,
        'total_labs_submitted': total_labs_submitted_all,
        'test_stats': test_stats,
        'test_grade_5_count': test_grade_5_count,
        'test_grade_4_count': test_grade_4_count,
        'test_grade_3_count': test_grade_3_count,
        'test_grade_2_count': test_grade_2_count,
        'test_grade_pending_count': test_grade_pending_count,
        'lab_grade_5_count': lab_grade_5_count,
        'lab_grade_4_count': lab_grade_4_count,
        'lab_grade_3_count': lab_grade_3_count,
        'lab_grade_2_count': lab_grade_2_count,
        'lab_grade_pending_count': lab_grade_pending_count,
        'theme_stats': list(theme_stats.items()),
        'lab_count': lab_count,
        'pending_submissions': pending_submissions,
        'groups_list': groups_list
    })
@login_required
def view_student_results(request):
    if request.user.profile.role != 'teacher':
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    # Получаем результаты тестов только тех учеников, которые закреплены за текущим учителем
    results = TestResult.objects.filter(
        student__profile__teacher=request.user
    ).order_related('student', 'test').order_by('-date_taken')

    return render(request, 'teacher_results.html', {
        'results': results
    })
@login_required
def student_detail(request, student_id):
    student = get_object_or_404(
        User,
        id=student_id,
        profile__role='student',
        profile__teacher=request.user
    )
    student_results = TestResult.objects.filter(student=student).order_by('-date_taken')

    return render(request, 'student_detail.html', {
        'student': student,
        'results': student_results
    })


@login_required
@teacher_required
def download_report(request):
    """Экспорт отчёта в Excel с 4 страницами: Тесты от учителя, Входное тестирование, Итоговое тестирование, Лабораторные работы"""

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Создаём новую книгу
    wb = Workbook()

    # Удаляем стандартный лист (создадим свои)
    wb.remove(wb.active)

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2f66ff", end_color="2f66ff", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Получаем всех учеников учителя
    students = User.objects.filter(
        profile__role='student',
        profile__teacher=request.user
    ).select_related('profile').order_by('profile__full_name')

    # ==================== 1. ТЕСТЫ ОТ УЧИТЕЛЯ ====================
    ws_teacher_tests = wb.create_sheet("Тесты от учителя")

    # Заголовки
    teacher_test_headers = ['№', 'Ученик', 'Группа', 'Курс', 'Название теста', 'Результат', '%', 'Оценка',
                            'Дата прохождения']
    for col, header in enumerate(teacher_test_headers, 1):
        cell = ws_teacher_tests.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 2
    for student in students:
        # Получаем все тесты от учителя для этого ученика
        teacher_tests_results = TestResult.objects.filter(
            user=student,
            test_type='teacher'
        ).select_related('teacher_test').order_by('-date_completed')

        for result in teacher_tests_results:
            ws_teacher_tests.cell(row=row, column=1, value=row - 1).border = border
            ws_teacher_tests.cell(row=row, column=2,
                                  value=student.profile.full_name or student.username).border = border
            ws_teacher_tests.cell(row=row, column=3, value=student.profile.group or '—').border = border
            ws_teacher_tests.cell(row=row, column=4, value=student.profile.course or '—').border = border
            ws_teacher_tests.cell(row=row, column=5,
                                  value=result.teacher_test.title if result.teacher_test else result.get_test_type_display()).border = border
            ws_teacher_tests.cell(row=row, column=6, value=f"{result.score}/{result.total_questions}").border = border
            ws_teacher_tests.cell(row=row, column=7, value=f"{result.percent}%").border = border
            ws_teacher_tests.cell(row=row, column=8,
                                  value=result.get_grade_display() if result.grade else '—').border = border
            ws_teacher_tests.cell(row=row, column=9,
                                  value=result.date_completed.strftime("%d.%m.%Y %H:%M")).border = border
            row += 1

    # Автоширина для страницы "Тесты от учителя"
    for col in range(1, 10):
        ws_teacher_tests.column_dimensions[get_column_letter(col)].auto_width = True

    # ==================== 2. ВХОДНОЕ ТЕСТИРОВАНИЕ ====================
    ws_start = wb.create_sheet("Входное тестирование")

    start_headers = ['№', 'Ученик', 'Группа', 'Курс', 'Результат', '%', 'Оценка', 'Дата прохождения',
                     'Комментарий учителя']
    for col, header in enumerate(start_headers, 1):
        cell = ws_start.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 2
    for student in students:
        start_result = TestResult.objects.filter(
            user=student,
            test_type='start'
        ).first()

        if start_result:
            ws_start.cell(row=row, column=1, value=row - 1).border = border
            ws_start.cell(row=row, column=2, value=student.profile.full_name or student.username).border = border
            ws_start.cell(row=row, column=3, value=student.profile.group or '—').border = border
            ws_start.cell(row=row, column=4, value=student.profile.course or '—').border = border
            ws_start.cell(row=row, column=5,
                          value=f"{start_result.score}/{start_result.total_questions}").border = border
            ws_start.cell(row=row, column=6, value=f"{start_result.percent}%").border = border
            ws_start.cell(row=row, column=7,
                          value=start_result.get_grade_display() if start_result.grade else '—').border = border
            ws_start.cell(row=row, column=8,
                          value=start_result.date_completed.strftime("%d.%m.%Y %H:%M")).border = border
            ws_start.cell(row=row, column=9, value=start_result.teacher_comment or '—').border = border
            row += 1
        else:
            ws_start.cell(row=row, column=1, value=row - 1).border = border
            ws_start.cell(row=row, column=2, value=student.profile.full_name or student.username).border = border
            ws_start.cell(row=row, column=3, value=student.profile.group or '—').border = border
            ws_start.cell(row=row, column=4, value=student.profile.course or '—').border = border
            ws_start.cell(row=row, column=5, value="Не пройден").border = border
            ws_start.cell(row=row, column=6, value="—").border = border
            ws_start.cell(row=row, column=7, value="—").border = border
            ws_start.cell(row=row, column=8, value="—").border = border
            ws_start.cell(row=row, column=9, value="—").border = border
            row += 1

    for col in range(1, 10):
        ws_start.column_dimensions[get_column_letter(col)].auto_width = True

    # ==================== 3. ИТОГОВОЕ ТЕСТИРОВАНИЕ ====================
    ws_final = wb.create_sheet("Итоговое тестирование")

    final_headers = ['№', 'Ученик', 'Группа', 'Курс', 'Результат', '%', 'Оценка', 'Дата прохождения',
                     'Комментарий учителя']
    for col, header in enumerate(final_headers, 1):
        cell = ws_final.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 2
    for student in students:
        final_result = TestResult.objects.filter(
            user=student,
            test_type='final'
        ).first()

        if final_result:
            ws_final.cell(row=row, column=1, value=row - 1).border = border
            ws_final.cell(row=row, column=2, value=student.profile.full_name or student.username).border = border
            ws_final.cell(row=row, column=3, value=student.profile.group or '—').border = border
            ws_final.cell(row=row, column=4, value=student.profile.course or '—').border = border
            ws_final.cell(row=row, column=5,
                          value=f"{final_result.score}/{final_result.total_questions}").border = border
            ws_final.cell(row=row, column=6, value=f"{final_result.percent}%").border = border
            ws_final.cell(row=row, column=7,
                          value=final_result.get_grade_display() if final_result.grade else '—').border = border
            ws_final.cell(row=row, column=8,
                          value=final_result.date_completed.strftime("%d.%m.%Y %H:%M")).border = border
            ws_final.cell(row=row, column=9, value=final_result.teacher_comment or '—').border = border
            row += 1
        else:
            ws_final.cell(row=row, column=1, value=row - 1).border = border
            ws_final.cell(row=row, column=2, value=student.profile.full_name or student.username).border = border
            ws_final.cell(row=row, column=3, value=student.profile.group or '—').border = border
            ws_final.cell(row=row, column=4, value=student.profile.course or '—').border = border
            ws_final.cell(row=row, column=5, value="Не пройден").border = border
            ws_final.cell(row=row, column=6, value="—").border = border
            ws_final.cell(row=row, column=7, value="—").border = border
            ws_final.cell(row=row, column=8, value="—").border = border
            ws_final.cell(row=row, column=9, value="—").border = border
            row += 1

    for col in range(1, 10):
        ws_final.column_dimensions[get_column_letter(col)].auto_width = True

    # ==================== 4. ЛАБОРАТОРНЫЕ РАБОТЫ ====================
    ws_labs = wb.create_sheet("Лабораторные работы")

    lab_headers = ['№', 'Ученик', 'Группа', 'Курс', 'Название работы', 'Статус', 'Оценка', 'Комментарий', 'Дата сдачи',
                   'Дата проверки']
    for col, header in enumerate(lab_headers, 1):
        cell = ws_labs.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 2
    for student in students:
        lab_submissions = LabSubmission.objects.filter(
            student=student,
            lab_work__created_by=request.user
        ).select_related('lab_work', 'graded_by__profile').order_by('-submitted_at')

        for submission in lab_submissions:
            status_display = {
                'submitted': 'Сдано',
                'under_review': 'На проверке',
                'graded': 'Проверено',
                'rejected': 'Отклонено'
            }.get(submission.status, submission.status)

            ws_labs.cell(row=row, column=1, value=row - 1).border = border
            ws_labs.cell(row=row, column=2, value=student.profile.full_name or student.username).border = border
            ws_labs.cell(row=row, column=3, value=student.profile.group or '—').border = border
            ws_labs.cell(row=row, column=4, value=student.profile.course or '—').border = border
            ws_labs.cell(row=row, column=5, value=submission.lab_work.title).border = border
            ws_labs.cell(row=row, column=6, value=status_display).border = border
            ws_labs.cell(row=row, column=7, value=submission.grade or '—').border = border
            ws_labs.cell(row=row, column=8, value=submission.comment or '—').border = border
            ws_labs.cell(row=row, column=9, value=submission.submitted_at.strftime("%d.%m.%Y %H:%M")).border = border
            ws_labs.cell(row=row, column=10, value=submission.graded_at.strftime(
                "%d.%m.%Y %H:%M") if submission.graded_at else '—').border = border
            row += 1

    for col in range(1, 11):
        ws_labs.column_dimensions[get_column_letter(col)].auto_width = True

    # Формируем имя файла
    now = timezone.localtime(timezone.now())
    filename = f"full_report_{now.strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Сохраняем в response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    # Например, для страницы "Тесты от учителя" - добавить строку с итогами
    if row > 2:
        # Пустая строка-разделитель
        ws_teacher_tests.cell(row=row, column=1, value="").border = border
        row += 1

        # Строка с итогами
        total_students_with_results = len(set(
            TestResult.objects.filter(
                user__profile__teacher=request.user,
                test_type='teacher'
            ).values_list('user_id', flat=True).distinct()
        ))

        summary_row = row
        ws_teacher_tests.cell(row=summary_row, column=2, value="ИТОГО:").font = Font(bold=True)
        ws_teacher_tests.cell(row=summary_row, column=5, value=f"Всего сдач: {row - 2}").font = Font(bold=True)
        ws_teacher_tests.cell(row=summary_row, column=6,
                              value=f"Учеников с результатами: {total_students_with_results}").font = Font(bold=True)

        # Объединяем ячейки для итогов
        ws_teacher_tests.merge_cells(f'A{summary_row}:A{summary_row + 1}')
        ws_teacher_tests.merge_cells(f'B{summary_row}:D{summary_row + 1}')
    wb.save(response)
    return response


@login_required
@teacher_required
def grade_test_result(request, result_id):
    result = get_object_or_404(TestResult, id=result_id)
    if result.user.profile.teacher != request.user:
        raise PermissionDenied("Нет прав на оценку этого теста")

    if request.method == 'POST':
        grade = request.POST.get('grade')
        comment = request.POST.get('comment', '')

        # Валидация оценки
        valid_grades = ['2', '3', '4', '5', 'н']
        if grade not in valid_grades:
            messages.error(request, 'Неверная оценка')
            return render(request, 'teacher/grade_test.html', {'result': result})

        result.grade = grade
        result.grade_comment = comment
        result.grade_date = timezone.now()
        result.save()

        messages.success(request, f'Оценка "{grade}" выставлена за тест!')
        return redirect('student_results')

    return render(request, 'teacher/grade_test.html', {'result': result})


@login_required
@teacher_required
def teacher_labs(request):
    labs = LabWork.objects.filter(created_by=request.user).order_by('-created_at')

    submissions = LabSubmission.objects.filter(
        lab_work__created_by=request.user
    ).select_related('student__profile', 'lab_work')

    # Фильтрация по проверенности через grade
    checked_filter = request.GET.get('checked')
    if checked_filter == 'true':
        submissions = submissions.filter(
            Q(grade__isnull=False) & ~Q(grade='')
        )
    elif checked_filter == 'false':
        submissions = submissions.filter(
            Q(grade__isnull=True) | Q(grade='')
        )

    # Фильтрация по ФИО студента
    search_name = request.GET.get('search_name', '').strip()
    if search_name:
        submissions = submissions.filter(
            student__profile__full_name__icontains=search_name
        )

    # НОВЫЕ ФИЛЬТРЫ: по группе и курсу
    group_filter = request.GET.get('group', '').strip()
    if group_filter:
        submissions = submissions.filter(student__profile__group=group_filter)

    course_filter = request.GET.get('course', '').strip()
    if course_filter:
        submissions = submissions.filter(student__profile__course=course_filter)

    submissions = submissions.order_by('-submitted_at')
    total_submissions = submissions.count()

    # Получаем уникальные группы и курсы для выпадающих списков
    students = User.objects.filter(profile__teacher=request.user, profile__role='student')
    available_groups = sorted(set(
        student.profile.group for student in students
        if student.profile and student.profile.group
    ))
    available_courses = sorted(set(
        student.profile.course for student in students
        if student.profile and student.profile.course
    ))

    # Добавляем информацию о группе и курсе для каждой сдачи
    for submission in submissions:
        profile = submission.student.profile
        submission.student_group = profile.group if profile and profile.group else '—'
        submission.student_course = profile.course if profile and profile.course else '—'

    return render(request, 'teacher/teacher_labs.html', {
        'labs': labs,
        'submissions': submissions,
        'total_submissions': total_submissions,
        'checked_filter': checked_filter,
        'search_name': search_name,
        'group_filter': group_filter,
        'course_filter': course_filter,
        'available_groups': available_groups,
        'available_courses': available_courses,
    })
@login_required
@teacher_required
def delete_lab_work(request, lab_id):
    """Удаление лабораторной работы (только для создавшего учителя)"""
    lab = get_object_or_404(LabWork, id=lab_id)

    # Проверяем, что текущий пользователь - создатель лабораторной
    if lab.created_by != request.user:
        messages.error(request, 'Вы не можете удалить чужую лабораторную работу')
        return redirect('teacher_labs')

    # Получаем название перед удалением для сообщения
    lab_title = lab.title

    # Удаляем лабораторную работу (все связанные submission удалятся каскадно)
    lab.delete()

    messages.success(request, f'Лабораторная работа "{lab_title}" успешно удалена')
    return redirect('teacher_labs')


@login_required
@student_required
def student_labs(request):
    teacher_labs = LabWork.objects.filter(
        created_by=request.user.profile.teacher
    ).filter(is_active=True)

    # Свои сдачи
    my_submissions = LabSubmission.objects.filter(
        student=request.user
    ).select_related('lab_work', 'graded_by__profile').order_by('-submitted_at')

    # Получаем ID всех лабораторных работ, которые уже сданы
    submitted_lab_ids = set(submission.lab_work_id for submission in my_submissions)

    return render(request, 'student/student_labs.html', {
        'labs': teacher_labs,
        'my_submissions': my_submissions,
        'submitted_lab_ids': submitted_lab_ids  # Новый контекст
    })
@login_required
@teacher_required
def create_lab_work(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        theme = request.POST.get('theme')
        docx_file = request.FILES.get('docx_file')
        pptx_file = request.FILES.get('pptx_file')  # новое поле для презентации

        if not all([title, description, theme, docx_file]):
            messages.error(request, 'Заполните все обязательные поля')
            return redirect('teacher_labs')

        lab = LabWork.objects.create(
            title=title,
            description=description,
            theme=theme,
            docx_file=docx_file,
            pptx_file=pptx_file,  # сохраняем pptx файл
            created_by=request.user,
            difficulty=request.POST.get('difficulty', 'medium')
        )
        students = User.objects.filter(profile__teacher=request.user, profile__role='student')
        notify_students_about_new_lab(request.user, students, lab)
        messages.success(request, f'Лабораторная "{title}" создана!')
        return redirect('lab_detail', lab_id=lab.id)

    return render(request, 'teacher/create_lab.html')

@login_required
def lab_view(request, lab_id):
    try:
        lab = LabWork.objects.get(id=lab_id, is_active=True)
        if lab.created_by != request.user.profile.teacher:
            messages.error(request, "У вас нет доступа к этой лабораторной работе")
            return redirect('student_labs')
    except LabWork.DoesNotExist:
        messages.error(request, "Запрашиваемая лабораторная работа не найдена или недоступна")
        return redirect('student_labs')

    existing_submission = lab.submissions.filter(
        student=request.user,
        lab_work=lab
    ).first()

    context = {
        'lab': lab,
        'existing_submission': existing_submission,
    }
    return render(request, 'student/lab.html', context)

@login_required
@teacher_required
def lab_detail(request, lab_id):
    try:
        lab = LabWork.objects.get(id=lab_id, created_by=request.user)
    except LabWork.DoesNotExist:
        messages.error(request, "Лабораторная работа не найдена или у вас нет доступа к ней")
        return redirect('teacher_labs')

    submissions = lab.submissions.select_related('student__profile').order_by('-submitted_at')

    if request.method == 'POST':
        submission_id = request.POST.get('submission_id')
        grade = request.POST.get('grade')
        comment = request.POST.get('comment', '')

        if submission_id and grade:
            submission = get_object_or_404(LabSubmission, id=submission_id, lab_work=lab)
            submission.grade = grade
            submission.comment = comment
            submission.status = 'graded'
            submission.graded_at = timezone.now()
            submission.graded_by = request.user
            submission.save()
            messages.success(request, 'Оценка сохранена!')
            return redirect('lab_detail', lab_id=lab.id)

    return render(request, 'teacher/lab_detail.html', {
        'lab': lab,
        'submissions': submissions,
    })


# views.py — обновленный вызов в submit_lab

@login_required
@student_required
def submit_lab(request, lab_id):
    """Сдача лабораторной работы"""
    lab = get_object_or_404(LabWork, id=lab_id, is_active=True)

    if request.method == 'POST':
        submitted_file = request.FILES.get('submitted_file')
        comment = request.POST.get('comment', '').strip()

        if submitted_file:
            existing_submission = LabSubmission.objects.filter(
                lab_work=lab,
                student=request.user
            ).first()

            if existing_submission:
                existing_submission.submitted_file = submitted_file
                existing_submission.comment = comment
                existing_submission.status = 'under_review'
                existing_submission.submitted_at = timezone.now()
                existing_submission.save()
                messages.success(request, 'Работа успешно перезаписана и отправлена на проверку!')
            else:
                submission = LabSubmission.objects.create(
                    lab_work=lab,
                    student=request.user,
                    submitted_file=submitted_file,
                    comment=comment,
                    status='under_review'
                )
                # ✅ ПЕРЕДАЁМ REQUEST
                notify_teacher_about_submission(
                    lab.created_by,
                    request.user,
                    lab,
                    submission.id,
                    request=request  # ← Добавлено!
                )
                messages.success(request, 'Работа сдана на проверку!')

            return redirect('student_labs')

    return render(request, 'student/lab.html', {'lab': lab})
@login_required
@teacher_required
def add_question(request):
    """Добавление нового вопроса в базу"""
    if request.method == 'POST':
        form = AddQuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.created_by = request.user
            question.save()
            messages.success(request, 'Вопрос успешно добавлен!')
            return redirect('manage_questions')
    else:
        form = AddQuestionForm()
    return render(request, 'teacher/add_question.html', {'form': form})


@login_required
@teacher_required
def manage_questions(request):
    """Просмотр всех вопросов, созданных учителем"""
    questions = TestQuestion.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, 'teacher/manage_questions.html', {'questions': questions})


# views.py - обновите функцию create_teacher_test

@login_required
@teacher_required
def create_teacher_test(request):
    students = UserProfile.objects.filter(
        role='student',
        teacher=request.user
    ).select_related('user')

    unique_groups = sorted(set(s.group for s in students if s and s.group))
    unique_courses = sorted(set(s.course for s in students if s and s.course))

    if request.method == 'POST':
        test_form = TeacherTestWithPersonalForm(request.POST, user=request.user)

        if test_form.is_valid():
            # Получаем выбранные вопросы из разных источников
            selected_question_ids = request.POST.getlist('existing_questions')
            selected_personal_ids = request.POST.getlist('personal_questions')

            # Считаем количество динамических вопросов
            dynamic_questions_count = 0
            i = 0
            dynamic_questions_data = []

            while f'question_form-{i}-question_text' in request.POST:
                question_text = request.POST.get(f'question_form-{i}-question_text', '').strip()
                if question_text:
                    dynamic_questions_count += 1
                    # Сохраняем данные динамического вопроса для последующего создания
                    dynamic_questions_data.append({
                        'question_text': question_text,
                        'option_a': request.POST.get(f'question_form-{i}-option1', ''),
                        'option_b': request.POST.get(f'question_form-{i}-option2', ''),
                        'option_c': request.POST.get(f'question_form-{i}-option3', ''),
                        'option_d': request.POST.get(f'question_form-{i}-option4', ''),
                        'correct_option': request.POST.get(f'question_form-{i}-correct_answer', ''),
                        'category': request.POST.get(f'question_form-{i}-category', ''),
                    })
                i += 1

            total_questions_count = len(selected_question_ids) + len(selected_personal_ids) + dynamic_questions_count

            # Проверяем, что есть хотя бы один вопрос
            if total_questions_count == 0:
                messages.error(request, 'Невозможно создать тест без вопросов! Добавьте хотя бы один вопрос.')
                context = {
                    'test_form': test_form,
                    'all_questions': TestQuestion.objects.all(),
                    'personal_questions': TeacherPersonalQuestion.objects.filter(teacher=request.user),
                    'categories': TestCategory.objects.all(),
                    'students': students,
                    'unique_groups': unique_groups,
                    'unique_courses': unique_courses,
                }
                return render(request, 'teacher/create_teacher_test.html', context)

            # Создаем тест
            test = test_form.save(commit=False)
            test.teacher = request.user
            test.save()

            # 1. Добавляем вопросы из общей базы
            if selected_question_ids:
                selected_questions = TestQuestion.objects.filter(id__in=selected_question_ids)
                for order, question in enumerate(selected_questions):
                    TeacherTestQuestion.objects.create(test=test, question=question, order=order)

            # 2. Добавляем существующие личные вопросы учителя
            if selected_personal_ids:
                selected_personal = TeacherPersonalQuestion.objects.filter(
                    id__in=selected_personal_ids,
                    teacher=request.user
                )
                for order, question in enumerate(selected_personal):
                    TeacherTestPersonalQuestion.objects.create(test=test, question=question, order=order)

            # 3. Добавляем НОВЫЕ вопросы - теперь они создаются как ЛИЧНЫЕ вопросы учителя
            for idx, q_data in enumerate(dynamic_questions_data):
                # Получаем категорию (если выбрана)
                category_text = q_data['category']
                if category_text:
                    try:
                        category_obj = TestCategory.objects.get(id=category_text)
                        category_name = category_obj.name
                    except (TestCategory.DoesNotExist, ValueError):
                        category_name = ""
                else:
                    category_name = ""

                # Преобразуем правильный ответ из 1/2/3/4 в a/b/c/d
                correct_option_raw = q_data['correct_option']
                correct_option_map = {'1': 'a', '2': 'b', '3': 'c', '4': 'd'}
                correct_option = correct_option_map.get(correct_option_raw, 'a')

                # СОЗДАЕМ ЛИЧНЫЙ ВОПРОС УЧИТЕЛЯ (не в общую базу)
                personal_question = TeacherPersonalQuestion.objects.create(
                    teacher=request.user,
                    question_text=q_data['question_text'],
                    option_a=q_data['option_a'],
                    option_b=q_data['option_b'],
                    option_c=q_data['option_c'],
                    option_d=q_data['option_d'],
                    correct_option=correct_option,
                    category=category_name if category_name else None
                )

                # Связываем личный вопрос с тестом
                TeacherTestPersonalQuestion.objects.create(test=test, question=personal_question, order=idx)

            # 4. Назначаем тест ученикам
            selected_student_ids = request.POST.getlist('selected_students')
            if selected_student_ids:
                selected_students = User.objects.filter(
                    id__in=selected_student_ids,
                    profile__teacher=request.user,
                    profile__role='student'
                )
                test.assigned_to.set(selected_students)
            else:
                # Если ни один ученик не выбран, назначаем всем
                all_students = User.objects.filter(
                    profile__teacher=request.user,
                    profile__role='student'
                )
                test.assigned_to.set(all_students)
                if all_students.exists():
                    messages.info(request, 'Тест назначен всем вашим ученикам')

            # 5. Отправляем уведомления ученикам
            for student in test.assigned_to.all():
                notify_student_about_new_test(request.user, student, test, request=request)

            messages.success(
                request,
                f'Тест "{test.title}" успешно создан с {total_questions_count} вопросами и назначен {test.assigned_to.count()} ученикам! '
                f'Новые вопросы сохранены в ваши личные вопросы.'
            )
            return redirect('teacher_manage_tests')
        else:
            # Если форма не валидна, выводим ошибки
            for field, errors in test_form.errors.items():
                for error in errors:
                    messages.error(request, f'Ошибка в поле {field}: {error}')

    else:
        test_form = TeacherTestWithPersonalForm(user=request.user)

    all_questions = TestQuestion.objects.all()
    personal_questions = TeacherPersonalQuestion.objects.filter(teacher=request.user)
    categories = TestCategory.objects.all()

    context = {
        'test_form': test_form,
        'all_questions': all_questions,
        'personal_questions': personal_questions,
        'categories': categories,
        'students': students,
        'unique_groups': unique_groups,
        'unique_courses': unique_courses,
    }
    return render(request, 'teacher/create_teacher_test.html', context)

@login_required
@teacher_required
def teacher_manage_tests(request):
    """Управление тестами от учителя"""
    tests = TeacherTest.objects.filter(teacher=request.user).prefetch_related('questions', 'personal_questions').order_by('-created_at')
    return render(request, 'teacher/manage_tests.html', {'tests': tests})


@login_required
@teacher_required
def teacher_test_detail(request, test_id):
    """Детали теста от учителя"""
    try:
        # Получаем тест
        test = TeacherTest.objects.get(id=test_id, teacher=request.user)

        # Явно загружаем вопросы
        test_questions = list(test.questions.all())
        test_personal_questions = list(test.personal_questions.all())

        # Для отладки (можно удалить после проверки)
        print(f"=== teacher_test_detail ===")
        print(f"Test: {test.title}")
        print(f"Common questions: {len(test_questions)}")
        print(f"Personal questions: {len(test_personal_questions)}")
        print(f"Total: {len(test_questions) + len(test_personal_questions)}")

    except TeacherTest.DoesNotExist:
        messages.error(request, "Тест не найден или у вас нет доступа к нему")
        return redirect('teacher_manage_tests')

    return render(request, 'teacher/test_detail.html', {
        'test': test,
        'test_questions': test_questions,
        'test_personal_questions': test_personal_questions,
        'total_questions': len(test_questions) + len(test_personal_questions)
    })

@login_required
@teacher_required
def delete_teacher_test(request, test_id):
    """Удаление теста учителя"""
    test = get_object_or_404(TeacherTest, id=test_id, teacher=request.user)

    if request.method == 'POST':
        test_title = test.title
        test.delete()
        messages.success(request, f'Тест "{test_title}" успешно удален')
        return redirect('teacher_manage_tests')

    # Если GET запрос, показываем подтверждение (будет через JavaScript)
    return redirect('teacher_manage_tests')


@login_required
@student_required
def student_teacher_tests(request):
    teacher = request.user.profile.teacher

    # Исправлено: убрали Q(assigned_to__exact='') и Q(assigned_to__isnull=True)
    # Тесты доступны если:
    # 1. Назначены конкретному ученику (assigned_to=студент)
    # 2. Не назначены никому (assigned_to__isnull=True) - это и есть "доступно всем"
    tests = TeacherTest.objects.filter(
        teacher=teacher,
        is_active=True
    ).filter(
        Q(assigned_to=request.user) | Q(assigned_to__isnull=True)
    ).distinct()

    # Получаем ID тестов, которые уже пройдены
    completed_test_ids = set(
        TestResult.objects.filter(
            user=request.user,
            test_type='teacher',
            teacher_test__isnull=False
        ).values_list('teacher_test_id', flat=True)
    )

    # Создаем словарь результатов для быстрого доступа
    results_map = {
        result.teacher_test_id: result
        for result in TestResult.objects.filter(
            user=request.user,
            test_type='teacher',
            teacher_test__isnull=False
        ).select_related('teacher_test')
    }

    for test in tests:
        test.is_completed = test.id in completed_test_ids
        test.result = results_map.get(test.id)
        test.total_questions = test.questions.count() + test.personal_questions.count()
    return render(request, 'student/teacher_tests.html', {'tests': tests})


# views.py - обновите функцию take_teacher_test

@login_required
@student_required
def take_teacher_test(request, test_id):
    try:
        test = TeacherTest.objects.get(id=test_id, is_active=True)
        if not test.assigned_to.filter(id=request.user.id).exists():
            messages.error(request, "Вам не назначен этот тест")
            return redirect('student_teacher_tests')
    except TeacherTest.DoesNotExist:
        messages.error(request, "Запрашиваемый тест не найден или недоступен")
        return redirect('student_teacher_tests')

    if request.method == 'POST':
        correct = 0
        total = test.questions.count() + test.personal_questions.count()
        answers = []

        # Обрабатываем общие вопросы
        for question in test.questions.all():
            user_answer = request.POST.get(f'q_{question.id}')
            if user_answer:
                user_answer = user_answer.strip().lower()
            else:
                user_answer = None

            is_correct = False
            if user_answer is not None:
                is_correct = (user_answer == question.correct_option)

            if is_correct:
                correct += 1

            answers.append({
                'question': question,
                'user_answer': user_answer,
                'correct_answer': question.correct_option,
                'is_correct': is_correct,
                'is_personal': False
            })

        # Обрабатываем личные вопросы
        for question in test.personal_questions.all():
            user_answer = request.POST.get(f'q_personal_{question.id}')
            if user_answer:
                user_answer = user_answer.strip().lower()
            else:
                user_answer = None

            is_correct = False
            if user_answer is not None:
                is_correct = (user_answer == question.correct_option)

            if is_correct:
                correct += 1

            answers.append({
                'question': question,
                'user_answer': user_answer,
                'correct_answer': question.correct_option,
                'is_correct': is_correct,
                'is_personal': True
            })

        percent = round((correct / total) * 100, 2) if total else 0

        result = TestResult.objects.create(
            user=request.user,
            test_type='teacher',
            score=correct,
            total_questions=total,
            percent=percent,
            correct_answers=correct,
            percentage=percent,
            teacher_test=test,
            category_results={'teacher_test': {'name': test.title, 'correct': correct, 'total': total}}
        )

        for answer in answers:
            TestAnswer.objects.create(
                result=result,
                question_id=answer['question'].id,
                question_text=answer['question'].question_text,
                user_answer=answer['user_answer'],
                user_answer_text=_get_answer_text_for_personal(answer['question'], answer['user_answer']) if answer['is_personal'] else _get_answer_text(answer['question'], answer['user_answer']),
                correct_answer=answer['correct_answer'],
                correct_answer_text=_get_answer_text_for_personal(answer['question'], answer['correct_answer']) if answer['is_personal'] else _get_answer_text(answer['question'], answer['correct_answer']),
                is_correct=answer['is_correct']
            )

        notify_teacher_about_test_completion(test.teacher, request.user, result, request=request)
        messages.success(request, f'Тест завершен! Ваш результат: {correct}/{total} ({percent}%)')
        return redirect('result_detail', pk=result.id)

    # Собираем все вопросы для отображения
    all_questions = []
    for q in test.questions.all():
        all_questions.append({
            'id': q.id,
            'question_text': q.question_text,
            'option_a': q.option_a,
            'option_b': q.option_b,
            'option_c': q.option_c,
            'option_d': q.option_d,
            'is_personal': False
        })
    for q in test.personal_questions.all():
        all_questions.append({
            'id': f'personal_{q.id}',
            'question_text': q.question_text,
            'option_a': q.option_a,
            'option_b': q.option_b,
            'option_c': q.option_c,
            'option_d': q.option_d,
            'is_personal': True
        })

    return render(request, 'student/take_teacher_test.html', {'test': test, 'questions': all_questions})


def _get_answer_text_for_personal(question, option_letter):
    """Получает текст ответа для личного вопроса"""
    if not option_letter:
        return ''
    mapping = {
        'a': question.option_a,
        'b': question.option_b,
        'c': question.option_c,
        'd': question.option_d,
    }
    return mapping.get(option_letter.lower(), '')


@login_required
@teacher_required
def teacher_lab_detail(request, lab_id):
    # Получаем работу или возвращаем 404
    lab = get_object_or_404(LabWork, id=lab_id)

    # Проверяем, что пользователь - учитель
    if not request.user.profile.role == 'teacher':
        return redirect('home')

    # Проверяем, что работа принадлежит студенту этого учителя
    if lab.student.profile.teacher != request.user:
        return redirect('teacher_labs')

    context = {
        'lab': lab,
        'student': lab.student,
        'student_profile': lab.student.profile,
    }
    return render(request, 'teacher/teacher_lab_detail.html', context)


@login_required
@teacher_required
def submission_detail(request, submission_id):
    try:
        submission = LabSubmission.objects.select_related(
            'student__profile',
            'lab_work',
            'graded_by__profile'
        ).get(
            id=submission_id,
            lab_work__created_by=request.user
        )
    except LabSubmission.DoesNotExist:
        messages.error(request, "Сданная работа не найдена или у вас нет доступа к ней")
        return redirect('teacher_labs')

    if request.method == 'POST':
        grade = request.POST.get('grade')
        comment = request.POST.get('comment', '').strip()

        valid_grades = ['5', '4', '3', '2', 'н']
        if grade not in valid_grades:
            messages.error(request, 'Выберите корректную оценку.')
            return redirect('submission_detail', submission_id=submission.id)

        submission.grade = grade
        submission.comment = comment
        submission.checked = True
        submission.status = 'graded'
        submission.graded_at = timezone.now()
        submission.graded_by = request.user
        submission.save()
        notify_student_about_lab_grade(request.user, submission.student, submission, request=request)
        messages.success(request, 'Оценка за лабораторную работу сохранена.')
        return redirect('submission_detail', submission_id=submission.id)
    student_profile = getattr(submission.student, 'profile', None)
    student_submissions_count = LabSubmission.objects.filter(
        student=submission.student
    ).count()

    # Определяем тип файла для правильного отображения в шаблоне
    filename = os.path.basename(submission.submitted_file.name) if submission.submitted_file else ''
    file_extension = os.path.splitext(filename)[1].lower()

    # Расширения, которые можно просмотреть в iframe/embed
    viewable_extensions = {'.pdf', '.jpg', '.jpeg', '.png', '.gif'}
    can_preview = file_extension in viewable_extensions

    context = {
        'submission': submission,
        'lab': submission.lab_work,
        'student': submission.student,
        'student_profile': student_profile,
        'student_full_name': student_profile.full_name if student_profile else submission.student.username,
        'student_username': submission.student.username,
        'student_course': student_profile.course if student_profile else '-',
        'student_submissions_count': student_submissions_count,
        'submitted_file_name': filename,
        'submitted_file_extension': file_extension,
        'can_preview': can_preview,
        'file_download_url': f'/teacher/lab/submission/{submission.id}/download/',
    }

    return render(request, 'teacher/students_lab_detail.html', context)

# Добавьте функцию экспорта в Excel:
@login_required
@teacher_required
def export_student_results_excel(request):
    """Экспорт результатов учеников в Excel"""
    student_id = request.GET.get('student_id')
    type_filter = request.GET.get('type_filter', '')
    grade_filter = request.GET.get('grade_filter', '')
    group_filter = request.GET.get('group_filter', '')  # НОВОЕ
    course_filter = request.GET.get('course_filter', '')  # НОВОЕ

    # Получаем данные с теми же фильтрами, что и на странице
    students_qs = User.objects.filter(
        profile__role='student',
        profile__teacher=request.user
    ).select_related('profile')

    test_results = TestResult.objects.filter(
        user__profile__teacher=request.user
    ).select_related('user', 'user__profile').order_by('-date_completed')

    lab_results = LabSubmission.objects.filter(
        student__profile__teacher=request.user
    ).select_related('student__profile', 'lab_work').order_by('-submitted_at')

    if student_id:
        test_results = test_results.filter(user_id=student_id)
        lab_results = lab_results.filter(student_id=student_id)

    # НОВАЯ ФИЛЬТРАЦИЯ
    if group_filter:
        test_results = test_results.filter(user__profile__group=group_filter)
        lab_results = lab_results.filter(student__profile__group=group_filter)

    if course_filter:
        test_results = test_results.filter(user__profile__course=course_filter)
        lab_results = lab_results.filter(student__profile__course=course_filter)

    if type_filter == 'test':
        lab_results = lab_results.none()
    elif type_filter == 'lab':
        test_results = test_results.none()

    if grade_filter:
        if grade_filter == 'pending':
            test_results = test_results.filter(grade__isnull=True)
            lab_results = lab_results.filter(grade__isnull=True)
        else:
            test_results = test_results.filter(grade=grade_filter)
            lab_results = lab_results.filter(grade=grade_filter)

    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты учеников"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2f66ff", end_color="2f66ff", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки (добавлены колонки Группа и Курс)
    headers = ['Тип', 'Ученик', 'Группа', 'Курс', 'Название', 'Результат', 'Процент', 'Оценка', 'Дата']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Заполняем данными
    row = 2

    # Тесты
    for result in test_results:
        profile = result.user.profile
        user_name = profile.full_name or result.user.username
        title = result.get_test_type_display()
        score = f"{result.score}/{result.total_questions}"
        percent = f"{result.percent}%" if result.percent else "—"
        grade = result.get_grade_display() if result.grade else "—"
        date = result.date_completed.strftime("%d.%m.%Y %H:%M")
        group = profile.group if profile and profile.group else "—"
        course = profile.course if profile and profile.course else "—"

        ws.cell(row=row, column=1, value="Тест").border = border
        ws.cell(row=row, column=2, value=user_name).border = border
        ws.cell(row=row, column=3, value=group).border = border
        ws.cell(row=row, column=4, value=course).border = border
        ws.cell(row=row, column=5, value=title).border = border
        ws.cell(row=row, column=6, value=score).border = border
        ws.cell(row=row, column=7, value=percent).border = border
        ws.cell(row=row, column=8, value=grade).border = border
        ws.cell(row=row, column=9, value=date).border = border
        row += 1

    # Лабораторные работы
    for submission in lab_results:
        profile = submission.student.profile
        user_name = profile.full_name or submission.student.username
        title = f"Лабораторная: {submission.lab_work.title}"
        score = submission.grade or "—"
        percent = "—"
        grade = submission.grade or "—"
        date = submission.submitted_at.strftime("%d.%m.%Y %H:%M")
        group = profile.group if profile and profile.group else "—"
        course = profile.course if profile and profile.course else "—"

        ws.cell(row=row, column=1, value="Лабораторная").border = border
        ws.cell(row=row, column=2, value=user_name).border = border
        ws.cell(row=row, column=3, value=group).border = border
        ws.cell(row=row, column=4, value=course).border = border
        ws.cell(row=row, column=5, value=title).border = border
        ws.cell(row=row, column=6, value=score).border = border
        ws.cell(row=row, column=7, value=percent).border = border
        ws.cell(row=row, column=8, value=grade).border = border
        ws.cell(row=row, column=9, value=date).border = border
        row += 1

    # Автоматическая ширина колонок
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].auto_width = True

    # Формируем имя файла
    filename = f"student_results_{timezone.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Сохраняем в response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@login_required
def get_notifications(request):
    notifications = Notification.objects.filter(recipient=request.user)[:20]
    html = render_to_string('includes/notifications.html', {'notifications': notifications}, request=request)
    return JsonResponse({'html': html})
@login_required
def mark_notification_read(request, notification_id):
    """Отметить уведомление как прочитанное"""
    notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    notification.is_read = True
    notification.save()
    return JsonResponse({'status': 'ok'})


@login_required
def mark_all_notifications_read(request):
    """Отметить все уведомления как прочитанные"""
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)

    # Если AJAX запрос, возвращаем JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('ajax'):
        return JsonResponse({'status': 'ok', 'count': 0})

    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def get_unread_notifications_count(request):
    """Получение количества непрочитанных уведомлений"""
    count = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return JsonResponse({'count': count})

@login_required
def delete_notification(request, notification_id):
    """Удалить уведомление"""
    notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    notification.delete()
    return JsonResponse({'status': 'ok'})

@login_required
def delete_all_notifications(request):
    """Удалить все уведомления пользователя"""
    Notification.objects.filter(recipient=request.user).delete()
    return JsonResponse({'status': 'ok'})


@login_required
@teacher_required
def teacher_students_list(request):
    """Страница со списком всех учеников учителя и статистикой по сданным/несданным работам"""
    students = User.objects.filter(
        profile__role='student',
        profile__teacher=request.user
    ).select_related('profile').order_by('profile__full_name')

    students_data = []
    total_labs_count_value = LabWork.objects.filter(created_by=request.user, is_active=True).count() or 1  # избегаем деления на 0

    for student in students:
        # ВСЕ пройденные тесты (всех типов)
        all_test_results = TestResult.objects.filter(user=student)
        completed_tests = all_test_results.count()

        # Сданные лабораторные
        submitted_labs = LabSubmission.objects.filter(
            student=student,
            lab_work__created_by=request.user
        ).count()

        # Общее количество доступных тестов = входной(1) + итоговый(1) + тесты учителя
        teacher_tests_count = TeacherTest.objects.filter(
            assigned_to=student,
            is_active=True
        ).count() or 0
        total_available_tests = 2 + teacher_tests_count  # 2 - входной и итоговый

        # Средний балл за все тесты
        test_scores = [r.grade for r in all_test_results if r.grade and r.grade != 0]
        avg_grade = round(sum(test_scores) / len(test_scores), 1) if test_scores else None
        teacher_tests_passed = TestResult.objects.filter(
            user=student,
            test_type='teacher'
        ).count()

        # Количество назначенных тестов от учителя
        teacher_tests_count = TeacherTest.objects.filter(
            assigned_to=student,
            is_active=True
        ).count()
        students_data.append({
            'student': student,
            'full_name': student.profile.full_name or student.username,
            'group': student.profile.group or '—',
            'course': student.profile.course or '—',
            'teacher_tests_passed': teacher_tests_passed,
            'teacher_tests_count': teacher_tests_count,
            'completed_tests': completed_tests,
            'submitted_labs': submitted_labs,
            'total_labs': total_labs_count_value,
            'total_available_tests': total_available_tests,
            'avg_grade': avg_grade,
        })

    total_students = len(students)
    total_test_assignments = sum(s['total_available_tests'] for s in students_data)
    total_test_completions = sum(s['completed_tests'] for s in students_data)
    total_lab_submissions = sum(s['submitted_labs'] for s in students_data)

    context = {
        'students_data': students_data,
        'total_students': total_students,
        'total_test_assignments': total_test_assignments,
        'total_test_completions': total_test_completions,
        'total_lab_submissions': total_lab_submissions,
    }
    return render(request, 'teacher/teacher_students_list.html', context)


@login_required
@teacher_required
def teacher_student_stats(request, student_id):
    """Детальная страница статистики по выбранному ученику"""
    student = get_object_or_404(
        User,
        id=student_id,
        profile__role='student',
        profile__teacher=request.user
    )

    profile = student.profile

    # ========== ВСЕ ТЕСТЫ ==========
    # Входной и итоговый тесты
    start_test_result = TestResult.objects.filter(user=student, test_type='start').first()
    final_test_result = TestResult.objects.filter(user=student, test_type='final').first()

    # Тесты от учителя
    assigned_tests = TeacherTest.objects.filter(
        assigned_to=student,
        is_active=True
    ).order_by('-created_at')

    completed_teacher_test_results = TestResult.objects.filter(
        user=student,
        test_type='teacher'
    ).select_related('teacher_test').order_by('-date_completed')

    results_by_teacher_test_id = {r.teacher_test_id: r for r in completed_teacher_test_results if r.teacher_test_id}

    # Собираем все тесты в один список
    all_tests_data = []

    # Входной тест
    all_tests_data.append({
        'title': 'Входное тестирование',
        'description': 'Тест для определения начального уровня знаний',
        'is_completed': start_test_result is not None,
        'result': start_test_result,
        'score': f"{start_test_result.score}/{start_test_result.total_questions}" if start_test_result else None,
        'percent': start_test_result.percent if start_test_result else None,
        'grade': start_test_result.get_grade_display() if start_test_result and start_test_result.grade else None,
        'completed_at': start_test_result.date_completed if start_test_result else None,
        'detail_url': f'/teacher/result/{start_test_result.id}/' if start_test_result else None,
    })

    # Итоговый тест
    all_tests_data.append({
        'title': 'Итоговое тестирование',
        'description': 'Финальный тест по всему курсу',
        'is_completed': final_test_result is not None,
        'result': final_test_result,
        'score': f"{final_test_result.score}/{final_test_result.total_questions}" if final_test_result else None,
        'percent': final_test_result.percent if final_test_result else None,
        'grade': final_test_result.get_grade_display() if final_test_result and final_test_result.grade else None,
        'completed_at': final_test_result.date_completed if final_test_result else None,
        'detail_url': f'/teacher/result/{final_test_result.id}/' if final_test_result else None,
    })

    # Тесты от учителя
    for test in assigned_tests:
        result = results_by_teacher_test_id.get(test.id)
        all_tests_data.append({
            'title': test.title,
            'description': test.description or 'Описание отсутствует',
            'is_completed': result is not None,
            'result': result,
            'score': f"{result.score}/{result.total_questions}" if result else None,
            'percent': result.percent if result else None,
            'grade': result.get_grade_display() if result and result.grade else None,
            'completed_at': result.date_completed if result else None,
            'detail_url': f'/teacher/result/{result.id}/' if result else None,
        })

    # Статистика по всем тестам
    total_tests_available = len(all_tests_data)
    total_tests_completed = len([t for t in all_tests_data if t['is_completed']])
    all_test_grades = [t['result'].grade for t in all_tests_data if
                       t['is_completed'] and t['result'] and t['result'].grade]
    all_tests_avg_grade = round(sum(all_test_grades) / len(all_test_grades), 1) if all_test_grades else None

    # Статистика только по тестам учителя
    tests_data = []
    for test in assigned_tests:
        result = results_by_teacher_test_id.get(test.id)
        tests_data.append({
            'test': test,
            'is_completed': result is not None,
            'result': result,
            'score': f"{result.score}/{result.total_questions}" if result else None,
            'percent': result.percent if result else None,
            'grade': result.get_grade_display() if result and result.grade else None,
            'completed_at': result.date_completed if result else None,
        })

    completed_tests_count = len([t for t in tests_data if t['is_completed']])
    not_completed_tests_count = len([t for t in tests_data if not t['is_completed']])
    test_grades = [t['result'].grade for t in tests_data if t['is_completed'] and t['result'].grade]
    test_avg_grade = round(sum(test_grades) / len(test_grades), 1) if test_grades else None

    # ========== ЛАБОРАТОРНЫЕ ==========
    all_labs = LabWork.objects.filter(
        created_by=request.user,
        is_active=True
    ).order_by('-created_at')

    submitted_labs = LabSubmission.objects.filter(
        student=student,
        lab_work__created_by=request.user
    ).select_related('lab_work', 'graded_by__profile')

    submitted_by_lab_id = {s.lab_work_id: s for s in submitted_labs}

    labs_data = []
    for lab in all_labs:
        submission = submitted_by_lab_id.get(lab.id)
        labs_data.append({
            'lab': lab,
            'is_submitted': submission is not None,
            'submission': submission,
            'grade': submission.grade if submission else None,
            'status': submission.get_status_display() if submission else None,
            'submitted_at': submission.submitted_at if submission else None,
            'comment': submission.comment if submission else None,
        })

    submitted_labs_count = len([l for l in labs_data if l['is_submitted']])
    not_submitted_labs_count = len([l for l in labs_data if not l['is_submitted']])

    # ========== ВСЕ РАБОТЫ (для сводной таблицы) ==========
    test_results_all = []

    # Все тесты (включая входной и итоговый)
    for test_data in all_tests_data:
        if test_data['is_completed'] and test_data['result']:
            result = test_data['result']
            test_results_all.append({
                'type': 'test',
                'type_name': 'Тестирование' if result.test_type in ['start', 'final'] else 'Тест от учителя',
                'title': test_data['title'],
                'score': test_data['score'],
                'percent': test_data['percent'],
                'grade': test_data['grade'],
                'grade_value': result.grade,
                'date': test_data['completed_at'],
                'detail_url': test_data['detail_url'],
            })

    lab_results_all = []
    for submission in submitted_labs:
        lab_results_all.append({
            'type': 'lab',
            'type_name': 'Лабораторная работа',
            'title': submission.lab_work.title,
            'score': submission.grade or "—",
            'percent': None,
            'grade': submission.grade,
            'grade_value': int(submission.grade) if submission.grade and submission.grade.isdigit() else None,
            'date': submission.submitted_at,
            'status': submission.get_status_display(),
            'detail_url': f'/teacher/lab/submission/{submission.id}/',
        })

    all_works = test_results_all + lab_results_all
    all_works.sort(key=lambda x: x['date'], reverse=True)

    context = {
        'student': student,
        'profile': profile,
        'full_name': profile.full_name or student.username,
        'group': profile.group or '—',
        'course': profile.course or '—',
        'email': student.email,

        # Все тесты
        'all_tests_data': all_tests_data,
        'total_tests_available': total_tests_available,
        'total_tests_completed': total_tests_completed,
        'all_tests_avg_grade': all_tests_avg_grade,

        # Тесты от учителя
        'tests_data': tests_data,
        'completed_tests_count': completed_tests_count,
        'not_completed_tests_count': not_completed_tests_count,
        'test_avg_grade': test_avg_grade,

        # Лабораторные
        'labs_data': labs_data,
        'submitted_labs_count': submitted_labs_count,
        'not_submitted_labs_count': not_submitted_labs_count,

        # Сводная таблица
        'all_works': all_works,
    }

    return render(request, 'teacher/teacher_student_stats.html', context)


@login_required
@teacher_required
def edit_test_assignment(request, test_id):
    """Редактирование назначения учеников для существующего теста"""
    from .utils import notify_student_about_new_test

    test = get_object_or_404(TeacherTest, id=test_id, teacher=request.user)

    # Получаем всех учеников учителя
    all_students = User.objects.filter(
        profile__role='student',
        profile__teacher=request.user
    ).select_related('profile')

    # Получаем старых назначенных учеников
    old_assigned_ids = set(test.assigned_to.values_list('id', flat=True))

    if request.method == 'POST':
        selected_student_ids = request.POST.getlist('selected_students')
        new_assigned_ids = set(int(id) for id in selected_student_ids if id)
        auto_assign = request.POST.get('auto_assign_new_students') == 'on'

        # Находим новых учеников
        new_student_ids = new_assigned_ids - old_assigned_ids

        # Обновляем назначение
        test.assigned_to.clear()
        if selected_student_ids:
            test.assigned_to.add(*selected_student_ids)

        test.auto_assign_new_students = auto_assign
        test.save()

        # Отправляем уведомления каждому новому ученику
        notification_count = 0
        for student_id in new_student_ids:
            try:
                student = User.objects.get(id=student_id, profile__teacher=request.user)
                if student != request.user:  # Не отправляем уведомление самому себе
                    notify_student_about_new_test(request.user, student, test)
                    notification_count += 1
                    print(f"✅ Уведомление отправлено: {student.profile.full_name}")
            except User.DoesNotExist:
                print(f"❌ Студент с ID {student_id} не найден")

        if notification_count > 0:
            messages.success(request,
                             f'Назначение теста обновлено! Уведомления отправлены {notification_count} ученикам.')
        else:
            messages.success(request, f'Назначение теста "{test.title}" обновлено!')

        return redirect('teacher_test_detail', test_id=test.id)

    context = {
        'test': test,
        'all_students': all_students,
        'assigned_ids': old_assigned_ids,
    }
    return render(request, 'teacher/edit_test_assignment.html', context)
@login_required
def serve_submission_file(request, submission_id):
    """Отдельный view для отдачи файла с правильной обработкой"""
    submission = get_object_or_404(
        LabSubmission,
        id=submission_id,
        lab_work__created_by=request.user
    )

    if not submission.submitted_file:
        return HttpResponse("Файл не найден", status=404)

    filename = os.path.basename(submission.submitted_file.name)
    return get_file_response(submission.submitted_file, filename)


def custom_404(request, exception=None):
    """Кастомная страница 404 с учетом роли пользователя"""
    context = {}

    # Добавляем информацию о роли для правильной навигации
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        context['user_role'] = request.user.profile.role
        context['user_is_authenticated'] = True
    else:
        context['user_is_authenticated'] = False

    # Убедитесь, что шаблон существует по этому пути
    return render(request, 'errors.html', context, status=404)


@login_required
def get_chat_messages(request, user_id):
    """API для получения новых сообщений (без перезагрузки страницы)"""
    other_user = get_object_or_404(User, id=user_id)

    # Проверяем, что это одногруппник
    user_profile = request.user.profile
    other_profile = other_user.profile

    if (user_profile.group != other_profile.group or
            user_profile.course != other_profile.course or
            user_profile.teacher != other_profile.teacher):
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)

    last_id = request.GET.get('last_id')

    messages = Message.objects.filter(
        sender__in=[request.user, other_user],
        recipient__in=[request.user, other_user]
    )

    if last_id and last_id != 'null' and last_id != '0':
        try:
            last_id_int = int(last_id)
            messages = messages.filter(id__gt=last_id_int)
        except ValueError:
            pass

    messages = messages.order_by('created_at')

    # Помечаем сообщения от другого пользователя как прочитанные
    Message.objects.filter(
        sender=other_user,
        recipient=request.user,
        is_read=False
    ).update(is_read=True)

    messages_data = []
    for msg in messages:
        messages_data.append({
            'id': msg.id,
            'content': msg.content,
            'created_at': msg.created_at.strftime("%H:%M %d.%m.%Y"),
            'sender_id': msg.sender_id,
        })

    return JsonResponse({'messages': messages_data})


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def send_teacher_student_file(request):
    """API для отправки файла между учителем и учеником"""
    try:
        recipient_id = request.POST.get('recipient_id')
        file = request.FILES.get('file')

        if not file:
            return JsonResponse({'error': 'Файл не выбран'}, status=400)

        recipient = get_object_or_404(User, id=recipient_id)
        current_user = request.user

        # Проверяем права
        is_teacher = current_user.profile.role == 'teacher'
        is_student = current_user.profile.role == 'student'

        if is_student and recipient != current_user.profile.teacher:
            return JsonResponse({'error': 'Вы можете отправлять файлы только своему преподавателю'}, status=403)
        elif is_teacher and recipient.profile.teacher != current_user:
            return JsonResponse({'error': 'Этот ученик не принадлежит вам'}, status=403)

        # Сохраняем файл
        file_name = file.name
        file_path = default_storage.save(f'chat_files/{current_user.id}_{recipient.id}_{file_name}',
                                         ContentFile(file.read()))

        # Создаем сообщение
        message = TeacherStudentMessage.objects.create(
            sender=current_user,
            recipient=recipient,
            message_type='file',
            file_attachment=file_path,
            file_name=file_name,
            content=f'📎 {file_name}'
        )

        # Отправляем уведомление
        from .utils import notify_about_teacher_message_with_file
        notify_about_teacher_message_with_file(recipient, current_user, file_name)

        sender_name = current_user.profile.full_name or current_user.username
        if is_teacher:
            sender_name = f"👩‍🏫 {sender_name}"
        else:
            sender_name = f"👨‍🎓 {sender_name}"

        return JsonResponse({
            'success': True,
            'message': {
                'id': message.id,
                'content': message.content,
                'message_type': 'file',
                'file_url': message.file_attachment.url,
                'file_name': file_name,
                'created_at': message.created_at.strftime("%H:%M %d.%m.%Y"),
                'sender_id': message.sender_id,
                'sender_name': sender_name
            }
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def send_teacher_student_voice(request):
    """API для отправки голосового сообщения"""
    try:
        recipient_id = request.POST.get('recipient_id')
        voice_file = request.FILES.get('voice')
        duration = request.POST.get('duration', 0)

        if not voice_file:
            return JsonResponse({'error': 'Голосовое сообщение не записано'}, status=400)

        recipient = get_object_or_404(User, id=recipient_id)
        current_user = request.user

        # Проверяем права (аналогично файлам)
        is_teacher = current_user.profile.role == 'teacher'
        is_student = current_user.profile.role == 'student'

        if is_student and recipient != current_user.profile.teacher:
            return JsonResponse({'error': 'Доступ запрещен'}, status=403)
        elif is_teacher and recipient.profile.teacher != current_user:
            return JsonResponse({'error': 'Доступ запрещен'}, status=403)

        # Сохраняем голосовое сообщение
        voice_path = default_storage.save(f'voice_messages/{current_user.id}_{recipient.id}_{voice_file.name}',
                                          ContentFile(voice_file.read()))

        message = TeacherStudentMessage.objects.create(
            sender=current_user,
            recipient=recipient,
            message_type='voice',
            voice_message=voice_path,
            voice_duration=int(duration),
            content='🎤 Голосовое сообщение'
        )

        sender_name = current_user.profile.full_name or current_user.username
        if is_teacher:
            sender_name = f"👩‍🏫 {sender_name}"
        else:
            sender_name = f"👨‍🎓 {sender_name}"

        return JsonResponse({
            'success': True,
            'message': {
                'id': message.id,
                'content': message.content,
                'message_type': 'voice',
                'voice_url': message.voice_message.url,
                'voice_duration': message.voice_duration,
                'created_at': message.created_at.strftime("%H:%M %d.%m.%Y"),
                'sender_id': message.sender_id,
                'sender_name': sender_name
            }
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def add_message_reaction(request):
    """API для добавления реакции на сообщение"""
    try:
        data = json.loads(request.body)
        message_id = data.get('message_id')
        reaction = data.get('reaction')

        message = get_object_or_404(TeacherStudentMessage, id=message_id)

        # Проверяем права доступа к сообщению
        current_user = request.user
        if current_user not in [message.sender, message.recipient]:
            return JsonResponse({'error': 'Нет доступа к этому сообщению'}, status=403)

        # Обновляем реакции
        reactions = message.reactions or {}
        user_id_str = str(current_user.id)

        if reaction and reactions.get(user_id_str) == reaction:
            # Убираем реакцию, если нажали ту же
            del reactions[user_id_str]
            print(f"Removed reaction {reaction} from user {user_id_str}")
        elif reaction:
            reactions[user_id_str] = reaction
            print(f"Added reaction {reaction} from user {user_id_str}")

        message.reactions = reactions
        message.save()

        # Возвращаем обновлённые реакции
        return JsonResponse({
            'success': True,
            'reactions': reactions
        })

    except Exception as e:
        print(f"Error in add_message_reaction: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@student_required
def student_test_results(request):
    """Страница со списком всех пройденных тестов студента"""
    user = request.user

    # Получаем все результаты тестов пользователя
    test_results = TestResult.objects.filter(user=user).order_by('-date_completed')

    # Добавляем дополнительную информацию для каждого теста
    for result in test_results:
        # Определяем URL для просмотра результата
        result.detail_url = f'/student/result/{result.id}/'

        # Определяем иконку в зависимости от типа
        if result.test_type == 'start':
            result.icon = '📥'
            result.type_name = 'Входное тестирование'
        elif result.test_type == 'final':
            result.icon = '📤'
            result.type_name = 'Итоговое тестирование'
        elif result.test_type == 'teacher':
            result.icon = '📝'
            result.type_name = result.teacher_test.title if result.teacher_test else 'Тест от учителя'
        else:
            result.icon = '📊'
            result.type_name = 'Тест'

        # Преобразуем оценку в читаемый формат
        if result.grade:
            if result.grade == 5:
                result.grade_display = '5 (Отлично)'
            elif result.grade == 4:
                result.grade_display = '4 (Хорошо)'
            elif result.grade == 3:
                result.grade_display = '3 (Удовлетворительно)'
            elif result.grade == 2:
                result.grade_display = '2 (Неудовлетворительно)'
            else:
                result.grade_display = str(result.grade)
        else:
            result.grade_display = 'На проверке'

    # Статистика
    total_tests = test_results.count()
    avg_score = test_results.aggregate(Avg('score'))['score__avg']
    avg_percent = test_results.aggregate(Avg('percent'))['percent__avg']

    # Распределение по типам
    start_count = test_results.filter(test_type='start').count()
    final_count = test_results.filter(test_type='final').count()
    teacher_count = test_results.filter(test_type='teacher').count()

    context = {
        'test_results': test_results,
        'total_tests': total_tests,
        'avg_score': round(avg_score, 1) if avg_score else 0,
        'avg_percent': round(avg_percent, 1) if avg_percent else 0,
        'start_count': start_count,
        'final_count': final_count,
        'teacher_count': teacher_count,
        'active_nav': 'test_results'
    }

    return render(request, 'student/test_results.html', context)


# views.py - добавьте новые функции

@login_required
@teacher_required
def manage_personal_questions(request):
    """Управление личными вопросами учителя"""
    questions = TeacherPersonalQuestion.objects.filter(teacher=request.user).order_by('-created_at')
    return render(request, 'teacher/manage_personal_questions.html', {'questions': questions})


@login_required
@teacher_required
def add_personal_question(request):
    """Добавление личного вопроса"""
    if request.method == 'POST':
        form = TeacherPersonalQuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.teacher = request.user
            question.save()
            messages.success(request, 'Личный вопрос успешно добавлен!')
            return redirect('manage_personal_questions')
    else:
        form = TeacherPersonalQuestionForm()

    return render(request, 'teacher/add_personal_question.html', {'form': form})


@login_required
@teacher_required
def delete_personal_question(request, question_id):
    """Удаление личного вопроса"""
    from django.db import transaction
    from .models import TeacherPersonalQuestion, TeacherTestPersonalQuestion

    question = get_object_or_404(TeacherPersonalQuestion, id=question_id, teacher=request.user)

    try:
        with transaction.atomic():
            # Удаляем связи с тестами
            TeacherTestPersonalQuestion.objects.filter(question=question).delete()
            # Удаляем вопрос
            question.delete()
        messages.success(request, 'Вопрос успешно удалён')
    except Exception as e:
        messages.error(request, f'Ошибка при удалении: {str(e)}')

    return redirect('manage_personal_questions')


@login_required
@teacher_required
def edit_personal_question(request, question_id):
    """Редактирование личного вопроса"""
    question = get_object_or_404(TeacherPersonalQuestion, id=question_id, teacher=request.user)

    if request.method == 'POST':
        form = TeacherPersonalQuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, 'Вопрос успешно обновлён!')
            return redirect('manage_personal_questions')
    else:
        form = TeacherPersonalQuestionForm(instance=question)

    return render(request, 'teacher/edit_personal_question.html', {'form': form, 'question': question})


# views.py - добавьте новую функцию

@login_required
@teacher_required
@require_http_methods(["POST"])
def delete_question_ajax(request, question_id):
    """AJAX удаление вопроса"""
    from django.db import transaction
    from .models import TestQuestion, TeacherTestQuestion, TeacherTestPersonalQuestion

    try:
        question = get_object_or_404(TestQuestion, id=question_id)

        # Проверка прав: только создатель вопроса может его удалить
        # Или суперпользователь
        if question.created_by != request.user and not request.user.is_superuser:
            return JsonResponse({
                'success': False,
                'error': 'Вы не можете удалить вопрос, созданный другим учителем'
            }, status=403)

        # Проверяем, используется ли вопрос в тестах
        used_in_tests = TeacherTestQuestion.objects.filter(question=question).count()

        with transaction.atomic():
            # Удаляем связи с тестами
            TeacherTestQuestion.objects.filter(question=question).delete()
            # Удаляем сам вопрос
            question.delete()

        message = f'Вопрос успешно удалён'
        if used_in_tests > 0:
            message = f'Вопрос успешно удалён из {used_in_tests} тестов'

        return JsonResponse({
            'success': True,
            'message': message,
            'deleted_id': question_id
        })

    except TestQuestion.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Вопрос не найден'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Ошибка при удалении: {str(e)}'
        }, status=500)


# views.py - добавьте функцию для AJAX удаления личных вопросов

@login_required
@teacher_required
@require_http_methods(["POST"])
def delete_personal_question_ajax(request, question_id):
    """AJAX удаление личного вопроса учителя"""
    from django.db import transaction
    from .models import TeacherPersonalQuestion, TeacherTestPersonalQuestion

    try:
        question = get_object_or_404(TeacherPersonalQuestion, id=question_id, teacher=request.user)

        with transaction.atomic():
            # Удаляем связи с тестами
            TeacherTestPersonalQuestion.objects.filter(question=question).delete()
            # Удаляем вопрос
            question.delete()

        return JsonResponse({
            'success': True,
            'message': 'Вопрос успешно удалён',
            'deleted_id': question_id
        })

    except TeacherPersonalQuestion.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Вопрос не найден'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Ошибка при удалении: {str(e)}'
        }, status=500)


@login_required
@teacher_required
@require_http_methods(["POST"])
def add_question_to_test(request, test_id):
    """Добавление вопроса в существующий тест"""
    import json
    from .models import TeacherTest, TeacherTestQuestion, TeacherTestPersonalQuestion, TeacherPersonalQuestion

    test = get_object_or_404(TeacherTest, id=test_id, teacher=request.user)

    try:
        data = json.loads(request.body)
        question_text = data.get('question_text')
        option_a = data.get('option_a')
        option_b = data.get('option_b')
        option_c = data.get('option_c')
        option_d = data.get('option_d')
        correct_option = data.get('correct_option')
        question_type = data.get('question_type', 'personal')

        if question_type == 'common':
            # Добавляем в общую базу
            category = TestCategory.objects.first()
            question = TestQuestion.objects.create(
                category=category,
                question_text=question_text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                correct_option=correct_option,
                created_by=request.user
            )
            TeacherTestQuestion.objects.create(test=test, question=question, order=test.questions.count())
        else:
            # Личный вопрос (только для этого теста)
            question = TeacherPersonalQuestion.objects.create(
                teacher=request.user,
                question_text=question_text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                correct_option=correct_option
            )
            TeacherTestPersonalQuestion.objects.create(test=test, question=question,
                                                       order=test.personal_questions.count())

        return JsonResponse({'success': True, 'question_id': question.id})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@teacher_required
@require_http_methods(["POST"])
def edit_question_in_test(request, test_id):
    """Редактирование вопроса в тесте"""
    import json
    from .models import TeacherTest, TestQuestion, TeacherPersonalQuestion, TeacherTestQuestion, \
        TeacherTestPersonalQuestion

    test = get_object_or_404(TeacherTest, id=test_id, teacher=request.user)

    try:
        data = json.loads(request.body)
        question_id = data.get('question_id')
        question_type = data.get('question_type')
        question_text = data.get('question_text')
        option_a = data.get('option_a')
        option_b = data.get('option_b')
        option_c = data.get('option_c')
        option_d = data.get('option_d')
        correct_option = data.get('correct_option')

        if question_type == 'common':
            question = get_object_or_404(TestQuestion, id=question_id)
            # Проверка прав: только создатель вопроса может его редактировать
            if question.created_by != request.user and not request.user.is_superuser:
                return JsonResponse({'success': False, 'error': 'Вы не можете редактировать этот вопрос'}, status=403)

            question.question_text = question_text
            question.option_a = option_a
            question.option_b = option_b
            question.option_c = option_c
            question.option_d = option_d
            question.correct_option = correct_option
            question.save()
        else:
            question = get_object_or_404(TeacherPersonalQuestion, id=question_id, teacher=request.user)
            question.question_text = question_text
            question.option_a = option_a
            question.option_b = option_b
            question.option_c = option_c
            question.option_d = option_d
            question.correct_option = correct_option
            question.save()

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
@login_required
@teacher_required
@require_http_methods(["POST"])
def remove_question_from_test(request, test_id):
    """Удаление вопроса из теста (без удаления из БД)"""
    import json
    from .models import TeacherTest, TeacherTestQuestion, TeacherTestPersonalQuestion

    test = get_object_or_404(TeacherTest, id=test_id, teacher=request.user)

    try:
        data = json.loads(request.body)
        question_id = data.get('question_id')
        question_type = data.get('question_type')

        if question_type == 'common':
            TeacherTestQuestion.objects.filter(test=test, question_id=question_id).delete()
        else:
            TeacherTestPersonalQuestion.objects.filter(test=test, question_id=question_id).delete()

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})